import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
import datetime
import calendar
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================================
# CONFIGURAÇÃO
# =========================================

st.set_page_config(
    page_title="Ranking Geral de Pontos",
    page_icon="icons/ranking-da-pagina.png",
    layout="wide"
)

st.title("📈 Ranking Geral de Pontos")

# =========================================
# LEITURA E TRATAMENTO DAS ABAS
# =========================================

if "dados_prod" not in st.session_state:
    st.warning("⚠️ Carregue os dados na página principal primeiro.")
    st.stop()

dados = st.session_state["dados_prod"]

try:
    prod = dados["Prod"].copy()
    gpon = dados["Gpon"].copy()
except KeyError as erro:
    st.error(f"❌ Aba não encontrada: {erro}")
    st.stop()

# Tratamento dos pontos individualmente antes de juntar
prod["Pontos"] = pd.to_numeric(prod["Pontos"], errors="coerce").fillna(0)
gpon["Pontos"] = pd.to_numeric(gpon["Pontos"], errors="coerce").fillna(0)

total_prod = prod["Pontos"].sum()
total_gpon = gpon["Pontos"].sum()
total_geral = total_prod + total_gpon

# =========================================
# JUNTAR DADOS
# =========================================

df = pd.concat([prod, gpon], ignore_index=True)

# =========================================
# FILTROS (SIDEBAR)
# =========================================

# Injeção de CSS para personalizar a aparência da Sidebar e dos widgets
st.html("""
    <style>
    /* 1. Altera a cor do texto "Filtros" no topo da Sidebar */
    .stSidebar h2 {
        color: #012869 !important; /* Azul escuro */
        font-size: 26px !important;
        font-weight: 700 !important;
    }

    /* 2. Altera a cor das etiquetas (labels) dos widgets na Sidebar */
    .stSidebar [data-testid="stWidgetLabel"] p {
        color: #000047 !important; /* Azul escuro */
        font-size: 16px !important;
        font-weight: 600 !important;
    }

    /* 3. Altera a cor de fundo e do texto das TAGS SELECIONADAS no multiselect */
    .stSidebar [data-baseweb="tag"] {
        background-color: #F37C04 !important; /* Fundo laranja */
        color: #FFFFFF !important; /* Texto branco */
        border-radius: 4px !important;
    }

    /* 4. Altera a cor do ícone de "X" para fechar a tag */
    .stSidebar [data-baseweb="tag"] svg {
        fill: #FFFFFF !important;
    }
    </style>
    """)

st.sidebar.header("Filtros")

# FILTRO PROJETO
if "Projeto" in df.columns:
    projetos_disponiveis = sorted(df["Projeto"].dropna().unique())
    projetos_selecionados = st.sidebar.multiselect(
        "Projeto", options=projetos_disponiveis, default=projetos_disponiveis
    )
    df = df[df["Projeto"].isin(projetos_selecionados)]

# FILTRO SUPERVISOR
if "Supervisor" in df.columns:
    supervisores_disponiveis = sorted(df["Supervisor"].dropna().unique())
    supervisores_selecionados = st.sidebar.multiselect(
        "Supervisor", options=supervisores_disponiveis, default=supervisores_disponiveis
    )
    df = df[df["Supervisor"].isin(supervisores_selecionados)]

# FILTRO EQUIPE
if "Nome Equipe" in df.columns:
    equipes_disponiveis = sorted(df["Nome Equipe"].dropna().unique())
    equipes_selecionadas = st.sidebar.multiselect(
        "Equipe", options=equipes_disponiveis, default=equipes_disponiveis
    )
    df = df[df["Nome Equipe"].isin(equipes_selecionadas)]

# =========================================
# KPIs
# =========================================

col1, col2, col3, col4 = st.columns(4)

# Totais corrigidos utilizando a variável unificada 'df' (já filtrada)
total_equipes = df["Nome Equipe"].nunique(
) if "Nome Equipe" in df.columns else 0
total_registros = len(df)
total_filtrado = df["Pontos"].sum() if "Pontos" in df.columns else 0

col1.metric("Total Pontos (Geral)", f"{total_geral:,.0f}")
col2.metric("Total Pontos (Filtrado)", f"{total_filtrado:,.0f}")
col3.metric("Total Equipes (Filtrado)", total_equipes)
col4.metric("Total Registros (Filtrado)", total_registros)

st.divider()

# =========================================
# INFORMAÇÕES DE ATUALIZAÇÃO
# =========================================

ultima_atualizacao = (
    df["Data Agendamento"].max() if "Data Agendamento" in df.columns else None
)

hoje = datetime.date.today()
ano_atual = hoje.year
mes_atual = hoje.month

_, ultimo_dia_num = calendar.monthrange(ano_atual, mes_atual)
ult_dia = datetime.date(ano_atual, mes_atual, ultimo_dia_num)

# CORREÇÃO: Convertendo as datas explicitamente para string/ISO padrão do numpy para evitar conflito de tipos
data_inicio_np = np.datetime64(
    ultima_atualizacao.date() if pd.notna(ultima_atualizacao) else hoje
)
data_fim_np = np.datetime64(ult_dia) + np.timedelta64(1, "D")

dias_faltantes = np.busday_count(
    data_inicio_np, data_fim_np, weekmask="1111110"  # Segunda a Sábado
)

# =========================================
# COLORINDO O DATAFRAME
# =========================================


def colorir_metas(valor):
    if valor >= 275 and valor < 300:
        return "background-color: #FFEB9C; color: #9C5700"  # Amarelo claro
    if valor >= 300 and valor < 400:
        return "background-color: #C6EFCE; color: #006100"  # Verde claro
    elif valor >= 400:
        return "background-color: #1F497D; color: #FFFFFF"  # Azul escuro
    return "background-color: #F2F2F2"


def colorir_projecao(valor):
    # Fundo cinza com letra branca
    return "background-color: grey; color: white; font-weight: bold"


def negrito(valor):
    return "font-weight: bold"  # Letra em negrito


# =========================================
# RANKING GERAL
# =========================================


ranking = pd.DataFrame()
ranking2 = pd.DataFrame()

if "Nome Equipe" in df.columns:
    # Agrupamento baseado nos dados filtrados atualizados
    ranking = (
        df.groupby(["CódAuxEquipe", "Nome Equipe", "Supervisor"])["Pontos"]
        .sum()
        .reset_index()
        .sort_values("Pontos", ascending=False)
    )

    ranking2 = (
        df.groupby(["CódAuxEquipe", "Nome Equipe", "Supervisor"])["Pontos"]
        .sum()
        .reset_index()
        .sort_values("Pontos", ascending=False)
    )

    # ===========================================================
    # CRIAÇÃO DE VARIÁVEIS DE DIAS TRABALHADOS E MÉDIA DE PONTOS
    # ===========================================================

    # Cálculo dos dias trabalhados por equipe
    if "Dias Trab Tecnico" in df.columns:
        max_por_tecnico = df.groupby("Nome Equipe")["Dias Trab Tecnico"].max()
        valores_mapeados = ranking["Nome Equipe"].map(max_por_tecnico)
        dias_trabalhados = valores_mapeados.fillna(0).astype(int)
    else:
        dias_trabalhados = 0

    # Média de pontos por dia trabalhado (Evita divisão por zero)
    media_pontos = ranking["Pontos"] / dias_trabalhados.replace(0, np.nan)

    # ================================================
    # INSERÇÃO DE COLUNAS DE POSIÇÃO, META E PROJEÇÃO
    # ================================================

    ranking.insert(0, "Posição", range(1, len(ranking) + 1))
    ranking.insert(5, "Meta | 300", ranking["Pontos"] - 300)
    ranking.insert(6, "Meta | 350", ranking["Pontos"] - 350)
    ranking.insert(7, "Meta | 375", ranking["Pontos"] - 375)
    ranking.insert(8, "Meta | 400", ranking["Pontos"] - 400)
    ranking.insert(
        9, "Projeção", ranking["Pontos"] + (media_pontos * dias_faltantes))

    ranking2.insert(0, "Posição", range(1, len(ranking) + 1))
    ranking2.insert(5, "Meta Dia | 300",
                    (ranking["Pontos"] - 300) / dias_faltantes)
    ranking2.insert(6, "Meta Dia | 350",
                    (ranking["Pontos"] - 350) / dias_faltantes)
    ranking2.insert(7, "Meta Dia | 375",
                    (ranking["Pontos"] - 375) / dias_faltantes)
    ranking2.insert(8, "Meta Dia | 400",
                    (ranking["Pontos"] - 400) / dias_faltantes)
    ranking2.insert(
        9, "Projeção", ranking["Pontos"] + (media_pontos * dias_faltantes))

    # =========================================
    # EXIBIÇÃO DO RANKING
    # =========================================

    st.subheader("📊 Pontos, Metas e Projeção")

    por_dia = st.toggle("Meta por Dia")

    if por_dia:
        st.dataframe(
            (
                ranking2.style.format(
                    {
                        "Pontos": "{:,.1f}",
                        "Meta Dia | 300": "{:,.1f}",
                        "Meta Dia | 350": "{:,.1f}",
                        "Meta Dia | 375": "{:,.1f}",
                        "Meta Dia | 400": "{:,.1f}",
                        "Projeção": "{:,.1f}",
                    }
                )
                .map(negrito, subset=["Pontos"])
                .map(colorir_metas, subset=["Pontos"])
                .map(colorir_projecao, subset=["Projeção"])
            ),
            use_container_width=True,
            height=500,
            hide_index=True,
        )
    else:
        st.dataframe(
            (
                ranking.style.format(
                    {
                        "Pontos": "{:,.1f}",
                        "Meta | 300": "{:,.1f}",
                        "Meta | 350": "{:,.1f}",
                        "Meta | 375": "{:,.1f}",
                        "Meta | 400": "{:,.1f}",
                        "Projeção": "{:,.1f}",
                    }
                )
                .map(negrito, subset=["Pontos"])
                .map(colorir_metas, subset=["Pontos"])
                .map(colorir_projecao, subset=["Projeção"])
            ),
            use_container_width=True,
            height=500,
            hide_index=True,
        )

    # =========================================
    # TOP 10 GRÁFICO
    # =========================================

    if not ranking.empty:
        top10 = ranking.head(10)

        fig = px.bar(
            top10,
            x="Nome Equipe",
            y="Pontos",
            text_auto=True,
            title="Top 10 Equipes por Pontos",
            color="Pontos",
            color_continuous_scale="Oryel",
        )

        fig.update_xaxes(title_text="")
        fig.update_yaxes(title_text="")

        st.plotly_chart(fig, use_container_width=True)
        st.write("")

else:
    st.warning(
        "⚠️ A coluna 'Nome Equipe' não foi encontrada nos dados para gerar o ranking."
    )

if pd.notna(ultima_atualizacao):
    st.info(
        f"***Última Atualização:*** {pd.to_datetime(ultima_atualizacao).strftime('%d/%m/%Y')}"
    )

# =========================================
# EXPORTAR EXCEL
# =========================================


def exportar_excel(dataframe):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Ranking")

        # Formatação do arquivo Excel
        wb = writer.book
        ws = writer.sheets['Ranking']

        # Definição de estilos
        cor_cabecalho = PatternFill(
            start_color="FF012869", end_color='FF012869', fill_type='solid')
        fonte_cabecalho = Font("Calibri", size=10.5,
                               bold=True, color='FFFFFFFF')
        fonte = Font("Calibri", size=10.5, color='FF000000')

        borda_fina = Border(
            left=Side(style='thin', color='FFD9D9D9'),
            right=Side(style='thin', color='FFD9D9D9'),
            top=Side(style='thin', color='FFD9D9D9'),
            bottom=Side(style='thin', color='FFD9D9D9')
        )

        fundo_prox_f1 = PatternFill(
            start_color='FFFFEB9C', end_color='FFFFEB9C', fill_type='solid')
        fonte_prox_f1 = Font("Calibri", size=10.5, bold=True, color="FF9C5700")

        fundo_f1 = PatternFill(start_color='FFC6EFCE',
                               end_color='FFC6EFCE', fill_type='solid')
        fonte_f1 = Font("Calibri", size=10.5, bold=True, color="FF006100")

        fundo_f2 = PatternFill(start_color='FF1F497D',
                               end_color='FF1F497D', fill_type='solid')
        fonte_f2 = Font("Calibri", size=10.5, bold=True, color="FFFFFFFF")

        fundo_proj = PatternFill(
            start_color='FF181818', end_color='FF181818', fill_type='solid')
        fonte_proj = Font("Calibri", size=10.5, bold=True, color="FFFFFFFF")

        # Laço único para aplicar bordas, cores e alinhamentos em todas as células
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for celula in row:
                celula.border = borda_fina

                if celula.row == 1:  # Primeira linha (Cabeçalho)
                    celula.fill = cor_cabecalho
                    celula.font = fonte_cabecalho
                    celula.alignment = Alignment(
                        horizontal='center', vertical='center')
                else:  # Linhas de dados
                    celula.alignment = Alignment(
                        horizontal='center', vertical='center')
                    celula.font = fonte

        # Autoajuste da largura das colunas baseado no conteúdo (evita textos cortados)
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Formatação dos critérios de Pontos
        for lin in range(2, ws.max_row + 1):
            celula = ws.cell(row=lin, column=5)
            celula.number_format = '#,##0.00'
            if celula.value >= 275 and celula.value < 300:
                celula.fill = fundo_prox_f1
                celula.font = fonte_prox_f1
            elif celula.value >= 300 and celula.value < 400:
                celula.fill = fundo_f1
                celula.font = fonte_f1
            elif celula.value >= 400:
                celula.fill = fundo_f2
                celula.font = fonte_f2

        # Formatação em número
        cols = [6, 7, 8, 9]
        for lin in range(2, ws.max_row + 1):
            for col in cols:
                celula = ws.cell(row=lin, column=col)
                celula.number_format = '#,##0.00'

        # Formatação da projeção
        for lin in range(2, ws.max_row + 1):
            celula = ws.cell(row=lin, column=10)
            celula.number_format = '#,##0.00'
            celula.fill = fundo_proj
            celula.font = fonte_proj

    return output.getvalue()

dados_excel = None
nome_arq = ""

if not ranking.empty and not por_dia:
    dados_excel = exportar_excel(ranking)
    nome_arq = "ranking_meta_geral.xlsx"
elif not ranking2.empty and por_dia:
    dados_excel = exportar_excel(ranking2)
    nome_arq = "ranking_meta_por_dia.xlsx"

if dados_excel is not None:
    st.download_button(
        label=f"📥 Baixar {nome_arq.replace('.xlsx', '').replace('_', ' ').title()}",
        data=dados_excel,
        file_name=nome_arq,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )
